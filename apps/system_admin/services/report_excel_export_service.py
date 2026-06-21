from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from datetime import date, datetime
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .excel_export_service import AdminExcelExportService


class AdminReportExcelExportService:
    """Build a presentation-ready workbook for admin analytics reports."""

    BRAND_COLOR = "F59E0B"
    BRAND_DARK = "0F172A"
    MUTED_COLOR = "64748B"
    LIGHT_FILL = "F8FAFC"
    BORDER_COLOR = "E2E8F0"
    WHITE = "FFFFFF"
    SERIES_LABELS = {
        "users": "Người dùng",
        "events": "Sự kiện",
        "registrations": "Đăng ký",
        "tickets": "Vé",
        "checkins": "Check-in",
        "support": "Hỗ trợ",
    }
    REPORT_LABELS = {
        "all": "Toàn bộ báo cáo",
        "overview": "Tổng quan điều hành",
        "time_series": "Xu hướng tăng trưởng",
        "status": "Phân bổ trạng thái",
        "categories": "Hiệu suất danh mục",
        "faculties": "Phân bổ theo khoa",
        "events": "Hiệu suất sự kiện",
        "support": "Vận hành hỗ trợ",
        "organizer_requests": "Yêu cầu organizer",
    }
    CONTENT_TYPE = AdminExcelExportService.DEFAULT_CONTENT_TYPE

    @classmethod
    def build_response(
        cls,
        *,
        filename: str,
        overview: Mapping[str, object],
        report_type: str,
    ) -> HttpResponse:
        safe_filename = AdminExcelExportService._sanitize_filename(filename)
        response = HttpResponse(
            cls.build_bytes(overview=overview, report_type=report_type),
            content_type=cls.CONTENT_TYPE,
        )
        response["Content-Disposition"] = f'attachment; filename="{safe_filename}"'
        return response

    @classmethod
    def build_bytes(cls, *, overview: Mapping[str, object], report_type: str) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)
        workbook.properties.title = cls.REPORT_LABELS.get(report_type, "Báo cáo quản trị UEvent")
        workbook.properties.subject = "Báo cáo và thống kê hệ thống UEvent"
        workbook.properties.creator = "UEvent Admin"

        if report_type in {"all", "overview"}:
            cls._build_overview_sheet(workbook, overview, report_type)
        if report_type in {"all", "time_series"}:
            cls._build_trends_sheet(workbook, overview, report_type)
        if report_type in {"all", "status"}:
            cls._build_status_sheet(workbook, overview, report_type)
        if report_type in {"all", "events", "categories"}:
            cls._build_events_sheet(workbook, overview, report_type)
        if report_type in {"all", "faculties"}:
            cls._build_audience_sheet(workbook, overview, report_type)
        if report_type in {"all", "support", "organizer_requests"}:
            cls._build_operations_sheet(workbook, overview, report_type)

        if not workbook.worksheets:
            cls._build_overview_sheet(workbook, overview, "overview")

        workbook.active = 0
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @classmethod
    def _build_overview_sheet(cls, workbook: Workbook, overview: Mapping[str, object], report_type: str) -> None:
        sheet = cls._create_sheet(workbook, "Tổng quan", overview, report_type)
        metrics = overview.get("metrics", [])
        row = cls._add_table(
            sheet,
            start_row=7,
            title="CHỈ SỐ ĐIỀU HÀNH",
            headers=["Chỉ số", "Giá trị", "Thông tin đối chiếu", "Mô tả"],
            rows=[
                [item.get("label", ""), item.get("value", 0), item.get("helper", ""), item.get("description", "")]
                for item in metrics
            ],
            table_name="OverviewMetrics",
            number_columns={2},
        )

        health_start = row + 3
        health = overview.get("system_health", [])
        health_end = cls._add_table(
            sheet,
            start_row=health_start,
            title="SỨC KHỎE HỆ THỐNG",
            headers=["Chỉ số", "Giá trị", "Tổng", "Điểm (%)"],
            rows=[[item.get("label", ""), item.get("value", 0), item.get("total", 0), item.get("score", 0)] for item in health],
            table_name="SystemHealth",
            number_columns={2, 3},
            percent_columns={4},
        )
        if health:
            sheet.conditional_formatting.add(
                f"D{health_start + 2}:D{health_end}",
                DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color=cls.BRAND_COLOR),
            )

        cls._add_table(
            sheet,
            start_row=health_end + 3,
            title="ĐIỂM CẦN CHÚ Ý",
            headers=["Nội dung", "Chi tiết", "Mức độ"],
            rows=[[item.get("title", ""), item.get("description", ""), cls._translate_status(item.get("severity", ""))] for item in overview.get("insights", [])],
            table_name="ReportInsights",
        )
        cls._set_widths(sheet, {"A": 28, "B": 18, "C": 35, "D": 48})

    @classmethod
    def _build_trends_sheet(cls, workbook: Workbook, overview: Mapping[str, object], report_type: str) -> None:
        sheet = cls._create_sheet(workbook, "Xu hướng", overview, report_type)
        series = overview.get("time_series", {})
        names = [name for name in cls.SERIES_LABELS if series.get(name)]
        periods = sorted({point.get("period", "") for name in names for point in series.get(name, [])})
        values = {name: {point.get("period", ""): point.get("count", 0) for point in series.get(name, [])} for name in names}
        rows = [[cls._parse_date(period), *[values[name].get(period, 0) for name in names]] for period in periods]
        end_row = cls._add_table(
            sheet,
            start_row=7,
            title="XU HƯỚNG THEO THỜI GIAN",
            headers=["Thời gian", *[cls.SERIES_LABELS[name] for name in names]],
            rows=rows,
            table_name="TrendData",
            number_columns=set(range(2, len(names) + 2)),
            date_columns={1},
        )
        if rows and names:
            chart = LineChart()
            chart.title = "Xu hướng hoạt động"
            chart.y_axis.title = "Số lượng"
            chart.x_axis.title = "Thời gian"
            chart.style = 13
            chart.height = 9
            chart.width = 18
            chart.add_data(Reference(sheet, min_col=2, max_col=len(names) + 1, min_row=8, max_row=end_row), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=1, min_row=9, max_row=end_row))
            chart.legend.position = "b"
            sheet.add_chart(chart, "I7")
        cls._set_widths(sheet, {"A": 16, **{chr(66 + index): 16 for index in range(len(names))}})

    @classmethod
    def _build_status_sheet(cls, workbook: Workbook, overview: Mapping[str, object], report_type: str) -> None:
        sheet = cls._create_sheet(workbook, "Trạng thái", overview, report_type)
        breakdowns = overview.get("breakdowns", {})
        labels = {
            "users_by_status": "NGƯỜI DÙNG",
            "events_by_status": "SỰ KIỆN",
            "registrations_by_status": "ĐĂNG KÝ",
            "tickets_by_status": "VÉ",
            "support_by_status": "HỖ TRỢ",
            "support_by_priority": "MỨC ƯU TIÊN HỖ TRỢ",
            "organizer_requests_by_status": "YÊU CẦU ORGANIZER",
        }
        row = 7
        for index, (key, title) in enumerate(labels.items(), start=1):
            items = breakdowns.get(key, [])
            row = cls._add_table(
                sheet,
                start_row=row,
                title=title,
                headers=["Trạng thái", "Số lượng", "Tỷ lệ (%)"],
                rows=[[cls._translate_status(item.get("label", "")), item.get("count", 0), item.get("percentage", 0)] for item in items],
                table_name=f"StatusGroup{index}",
                number_columns={2},
                percent_columns={3},
            ) + 3
        cls._set_widths(sheet, {"A": 30, "B": 16, "C": 16})

    @classmethod
    def _build_events_sheet(cls, workbook: Workbook, overview: Mapping[str, object], report_type: str) -> None:
        sheet = cls._create_sheet(workbook, "Sự kiện", overview, report_type)
        row = 7
        if report_type in {"all", "categories"}:
            categories = overview.get("category_performance", [])
            end_row = cls._add_table(
                sheet,
                start_row=row,
                title="HIỆU SUẤT DANH MỤC",
                headers=["Danh mục", "Sự kiện", "Lượt đăng ký"],
                rows=[[item.get("label", ""), item.get("events_count", 0), item.get("registration_count", 0)] for item in categories],
                table_name="CategoryPerformance",
                number_columns={2, 3},
            )
            if categories:
                cls._add_bar_chart(sheet, title="Đăng ký theo danh mục", data_col=3, category_col=1, header_row=row + 1, end_row=end_row, anchor="E7")
            row = max(end_row + 3, 25)

        if report_type in {"all", "events"}:
            events = overview.get("top_events", [])
            cls._add_table(
                sheet,
                start_row=row,
                title="TOP SỰ KIỆN",
                headers=["Sự kiện", "Trạng thái", "Danh mục", "Sức chứa", "Đăng ký", "Check-in", "Tỷ lệ check-in (%)", "Lấp đầy (%)"],
                rows=[
                    [
                        item.get("title", ""), cls._translate_status(item.get("status", "")), item.get("category", ""),
                        item.get("max_capacity", 0), item.get("registration_count", 0), item.get("checkin_count", 0),
                        item.get("checkin_rate", 0), item.get("capacity_rate", 0),
                    ]
                    for item in events
                ],
                table_name="TopEvents",
                number_columns={4, 5, 6},
                percent_columns={7, 8},
            )
        cls._set_widths(sheet, {"A": 38, "B": 18, "C": 24, "D": 14, "E": 14, "F": 14, "G": 20, "H": 16})

    @classmethod
    def _build_audience_sheet(cls, workbook: Workbook, overview: Mapping[str, object], report_type: str) -> None:
        sheet = cls._create_sheet(workbook, "Người dùng", overview, report_type)
        faculties = overview.get("faculty_distribution", [])
        end_row = cls._add_table(
            sheet,
            start_row=7,
            title="PHÂN BỔ NGƯỜI THAM GIA THEO KHOA",
            headers=["Khoa", "Lượt đăng ký"],
            rows=[[item.get("label", ""), item.get("count", 0)] for item in faculties],
            table_name="FacultyDistribution",
            number_columns={2},
        )
        if faculties:
            cls._add_bar_chart(sheet, title="Lượt đăng ký theo khoa", data_col=2, category_col=1, header_row=8, end_row=end_row, anchor="D7")
        cls._set_widths(sheet, {"A": 36, "B": 18})

    @classmethod
    def _build_operations_sheet(cls, workbook: Workbook, overview: Mapping[str, object], report_type: str) -> None:
        sheet = cls._create_sheet(workbook, "Vận hành", overview, report_type)
        row = 7
        breakdowns = overview.get("breakdowns", {})
        if report_type in {"all", "support"}:
            for index, (key, title) in enumerate((("support_by_status", "HỖ TRỢ THEO TRẠNG THÁI"), ("support_by_priority", "HỖ TRỢ THEO ƯU TIÊN")), start=1):
                items = breakdowns.get(key, [])
                row = cls._add_table(
                    sheet, start_row=row, title=title, headers=["Nhóm", "Số lượng", "Tỷ lệ (%)"],
                    rows=[[cls._translate_status(item.get("label", "")), item.get("count", 0), item.get("percentage", 0)] for item in items],
                    table_name=f"SupportGroup{index}", number_columns={2}, percent_columns={3},
                ) + 3

        if report_type in {"all", "organizer_requests"}:
            summary = overview.get("organizer_request_summary", {})
            cls._add_table(
                sheet, start_row=row, title="YÊU CẦU TRỞ THÀNH ORGANIZER", headers=["Chỉ số", "Giá trị"],
                rows=[
                    ["Tổng yêu cầu", summary.get("total", 0)], ["Chờ duyệt", summary.get("pending", 0)],
                    ["Đã duyệt", summary.get("approved", 0)], ["Từ chối", summary.get("rejected", 0)],
                    ["Tỷ lệ duyệt (%)", summary.get("approval_rate", 0)],
                ],
                table_name="OrganizerRequests", number_columns={2},
            )
        cls._set_widths(sheet, {"A": 34, "B": 18, "C": 18})

    @classmethod
    def _create_sheet(cls, workbook: Workbook, title: str, overview: Mapping[str, object], report_type: str) -> Worksheet:
        sheet = workbook.create_sheet(title=title)
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.tabColor = cls.BRAND_COLOR
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.merge_cells("A1:H1")
        sheet["A1"] = "UEVENT | BÁO CÁO & THỐNG KÊ"
        sheet["A1"].font = Font(size=18, bold=True, color=cls.WHITE)
        sheet["A1"].fill = PatternFill("solid", fgColor=cls.BRAND_DARK)
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 34
        filters = overview.get("filters", {})
        metadata = [
            ("Phạm vi", cls.REPORT_LABELS.get(report_type, report_type)),
            ("Kỳ báo cáo", f"{filters.get('from_date', '')} - {filters.get('to_date', '')}"),
            ("Nhóm dữ liệu", cls._group_label(filters.get("group_by", "day"))),
            ("Thời điểm tạo", cls._format_datetime(overview.get("generated_at"))),
        ]
        for row, (label, value) in enumerate(metadata, start=2):
            sheet.cell(row, 1, label).font = Font(bold=True, color=cls.MUTED_COLOR)
            sheet.cell(row, 2, value).font = Font(color=cls.BRAND_DARK)
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        return sheet

    @classmethod
    def _add_table(
        cls,
        sheet: Worksheet,
        *,
        start_row: int,
        title: str,
        headers: Sequence[str],
        rows: Iterable[Sequence[object]],
        table_name: str,
        number_columns: set[int] | None = None,
        percent_columns: set[int] | None = None,
        date_columns: set[int] | None = None,
    ) -> int:
        data_rows = list(rows)
        end_column = len(headers)
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_column)
        title_cell = sheet.cell(start_row, 1, title)
        title_cell.font = Font(size=11, bold=True, color=cls.WHITE)
        title_cell.fill = PatternFill("solid", fgColor=cls.BRAND_DARK)
        title_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[start_row].height = 24

        header_row = start_row + 1
        thin_border = Border(bottom=Side(style="thin", color=cls.BORDER_COLOR))
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(header_row, column, header)
            cell.font = Font(bold=True, color=cls.BRAND_DARK)
            cell.fill = PatternFill("solid", fgColor="FEF3C7")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

        for row_offset, values in enumerate(data_rows, start=1):
            row_number = header_row + row_offset
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row_number, column, value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                if column in (number_columns or set()):
                    cell.number_format = "#,##0"
                if column in (percent_columns or set()):
                    cell.number_format = '0.00"%"'
                if column in (date_columns or set()):
                    cell.number_format = "dd/mm/yyyy"

        end_row = header_row + max(len(data_rows), 1)
        if data_rows:
            table = Table(displayName=table_name, ref=f"A{header_row}:{cls._column_letter(end_column)}{end_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
            sheet.add_table(table)
        return end_row

    @classmethod
    def _add_bar_chart(cls, sheet: Worksheet, *, title: str, data_col: int, category_col: int, header_row: int, end_row: int, anchor: str) -> None:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = title
        chart.height = 8
        chart.width = 15
        chart.legend = None
        chart.add_data(Reference(sheet, min_col=data_col, min_row=header_row, max_row=end_row), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=category_col, min_row=header_row + 1, max_row=end_row))
        sheet.add_chart(chart, anchor)

    @staticmethod
    def _set_widths(sheet: Worksheet, widths: Mapping[str, float]) -> None:
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

    @staticmethod
    def _column_letter(column: int) -> str:
        result = ""
        while column:
            column, remainder = divmod(column - 1, 26)
            result = chr(65 + remainder) + result
        return result

    @staticmethod
    def _parse_date(value: object) -> object:
        if not value:
            return ""
        try:
            return date.fromisoformat(str(value))
        except ValueError:
            return value

    @staticmethod
    def _format_datetime(value: object) -> str:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y %H:%M")
        return str(value or "")

    @staticmethod
    def _group_label(value: object) -> str:
        return {"day": "Theo ngày", "week": "Theo tuần", "month": "Theo tháng"}.get(str(value), str(value))

    @staticmethod
    def _translate_status(value: object) -> str:
        labels = {
            "pending": "Chờ xử lý", "approved": "Đã duyệt", "rejected": "Từ chối",
            "active": "Hoạt động", "banned": "Đình chỉ", "draft": "Nháp",
            "published": "Đã xuất bản", "cancelled": "Đã hủy", "completed": "Đã hoàn thành",
            "open": "Mở", "in_progress": "Đang xử lý", "resolved": "Đã giải quyết", "closed": "Đóng",
            "issued": "Đã cấp", "used": "Đã dùng", "expired": "Hết hạn", "revoked": "Thu hồi",
            "registered": "Đã đăng ký", "checked_in": "Đã check-in", "waitlisted": "Danh sách chờ",
            "success": "Tích cực", "warning": "Cần chú ý", "info": "Thông tin",
            "low": "Thấp", "medium": "Trung bình", "high": "Cao", "urgent": "Khẩn cấp",
        }
        return labels.get(str(value).lower(), str(value))
