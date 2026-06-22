from datetime import datetime, timezone
from io import BytesIO

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.system_admin.services.report_excel_export_service import AdminReportExcelExportService


class AdminReportExcelExportServiceTests(SimpleTestCase):
    def setUp(self):
        self.overview = {
            "generated_at": datetime(2026, 6, 21, 9, 30, tzinfo=timezone.utc),
            "filters": {"from_date": "2026-06-01", "to_date": "2026-06-21", "group_by": "day"},
            "metrics": [
                {"label": "Người dùng mới", "value": 24, "helper": "120 tài khoản", "description": "Tạo trong kỳ"},
            ],
            "time_series": {
                "users": [{"period": "2026-06-01", "count": 4}, {"period": "2026-06-02", "count": 6}],
                "events": [{"period": "2026-06-01", "count": 2}],
            },
            "breakdowns": {
                "users_by_status": [{"label": "active", "count": 20, "percentage": 83.33}],
                "events_by_status": [{"label": "approved", "count": 5, "percentage": 100}],
                "registrations_by_status": [],
                "tickets_by_status": [],
                "support_by_status": [{"label": "open", "count": 3, "percentage": 60}],
                "support_by_priority": [{"label": "high", "count": 2, "percentage": 40}],
                "organizer_requests_by_status": [{"label": "pending", "count": 2, "percentage": 50}],
            },
            "category_performance": [{"label": "Học thuật", "events_count": 3, "registration_count": 80}],
            "faculty_distribution": [{"label": "CNTT", "count": 45}],
            "top_events": [
                {
                    "title": "Ngày hội công nghệ", "status": "approved", "category": "Học thuật",
                    "max_capacity": 100, "registration_count": 80, "checkin_count": 60,
                    "checkin_rate": 75, "capacity_rate": 80,
                }
            ],
            "organizer_request_summary": {"total": 4, "pending": 2, "approved": 1, "rejected": 1, "approval_rate": 50},
            "system_health": [{"label": "Hiệu quả sử dụng vé", "value": 60, "total": 80, "score": 75}],
            "insights": [{"title": "Tỷ lệ check-in", "description": "Đạt 75% trong kỳ.", "severity": "success"}],
        }

    def test_all_report_builds_structured_multi_sheet_workbook(self):
        content = AdminReportExcelExportService.build_bytes(overview=self.overview, report_type="all")
        workbook = load_workbook(BytesIO(content))

        self.assertEqual(
            workbook.sheetnames,
            ["Tổng quan", "Xu hướng", "Trạng thái", "Sự kiện", "Người dùng", "Vận hành"],
        )
        self.assertEqual(workbook["Tổng quan"]["A1"].value, "UEVENT | BÁO CÁO & THỐNG KÊ")
        self.assertEqual(workbook["Tổng quan"]["B3"].value, "2026-06-01 - 2026-06-21")
        self.assertEqual(workbook["Tổng quan"]["B9"].value, 24)
        self.assertIsNone(workbook["Tổng quan"].freeze_panes)
        self.assertTrue(all(sheet.freeze_panes is None for sheet in workbook.worksheets))
        self.assertGreaterEqual(len(workbook["Xu hướng"]._charts), 1)
        self.assertGreaterEqual(len(workbook["Sự kiện"]._charts), 1)

    def test_specialized_export_contains_only_relevant_sheet(self):
        content = AdminReportExcelExportService.build_bytes(overview=self.overview, report_type="faculties")
        workbook = load_workbook(BytesIO(content))

        self.assertEqual(workbook.sheetnames, ["Người dùng"])
        self.assertEqual(workbook["Người dùng"]["A8"].value, "Khoa")
        self.assertEqual(workbook["Người dùng"]["B9"].value, 45)
        self.assertEqual(workbook["Người dùng"]["B9"].number_format, "#,##0")

    def test_response_keeps_xlsx_download_contract(self):
        response = AdminReportExcelExportService.build_response(
            filename="admin_reports_all_20260621.xlsx",
            overview=self.overview,
            report_type="all",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], AdminReportExcelExportService.CONTENT_TYPE)
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="admin_reports_all_20260621.xlsx"')
